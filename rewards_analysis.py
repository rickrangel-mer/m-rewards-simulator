import os
from pathlib import Path

import openpyxl
import pandas as pd
from pyathena import connect

NUM_MONTHS = 6
OUTPUT_DIR = Path(__file__).parent
EXCEL_PATH = OUTPUT_DIR / "M-rewards-cocacola.xlsx"
CACHE_PATH = OUTPUT_DIR / "rewards_raw_data.csv"

ATHENA_REGION = "us-west-2"
ATHENA_S3_STAGING = "s3://mercaso-data-platform-prod/athena/sql/"


# ---------------------------------------------------------------------------
# I/O: Excel loading
# ---------------------------------------------------------------------------

def load_sku_mapping(excel_path=EXCEL_PATH):
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    ims_sheet = wb["ims"]
    sku_to_title = {}
    for row in ims_sheet.iter_rows(min_row=2, values_only=True):
        sku_to_title[row[3]] = row[4]  # col D = SKU, col E = Product Title - Long

    ws = wb["WOS Ranked"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = row[2]   # col C
        if not sku:
            continue
        records.append({
            "sku": sku,
            "product_title": sku_to_title.get(sku, row[3]),  # fallback to col D
            "category": row[1],   # col B
            "rewards_match": row[11],  # col L
            "points": row[10],  # col K
        })
    wb.close()

    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# I/O: Athena query with cache
# ---------------------------------------------------------------------------

def fetch_order_data(sku_list, cache_path=CACHE_PATH):
    if cache_path.exists():
        print(f"Loading cached data from {cache_path}")
        df = pd.read_csv(cache_path, parse_dates=["order_date"])
        return df

    print("Querying Athena (no cache found)...")
    sku_values = ", ".join(f"'{s}'" for s in sku_list)
    query = f"""
    SELECT
        o.store_id,
        li.sku,
        DATE(li.order_item_created_at) AS order_date,
        SUM(li.initial_quantity)       AS total_quantity
    FROM dwm.dwm_trade_line_item_detail_full li
    JOIN dwm.dwm_trade_order_detail_full o
        ON li.order_id = o.order_id
       AND o.dt = (SELECT MAX(dt) FROM dwm.dwm_trade_order_detail_full)
    WHERE li.dt = (SELECT MAX(dt) FROM dwm.dwm_trade_line_item_detail_full)
      AND li.order_item_created_at >= TIMESTAMP '2026-01-01'
      AND li.sku IN ({sku_values})
    GROUP BY o.store_id, li.sku, DATE(li.order_item_created_at)
    ORDER BY o.store_id, li.sku, DATE(li.order_item_created_at)
    """

    conn = connect(
        s3_staging_dir=ATHENA_S3_STAGING,
        region_name=ATHENA_REGION,
    )
    df = pd.read_sql(query, conn)
    df["order_date"] = pd.to_datetime(df["order_date"])

    df.to_csv(cache_path, index=False)
    print(f"Cached {len(df):,} rows to {cache_path}")
    return df


# ---------------------------------------------------------------------------
# Computation: cohort assignment
# ---------------------------------------------------------------------------

def assign_store_cohorts(order_data: pd.DataFrame, rewards_skus: set) -> dict:
    stores_with_rewards = set(
        order_data.loc[order_data["sku"].isin(rewards_skus), "store_id"]
    )
    all_stores = set(order_data["store_id"])
    return {
        store: "rewards-eligible" if store in stores_with_rewards else "coca-cola-only"
        for store in all_stores
    }


# ---------------------------------------------------------------------------
# Computation: SKU-level metrics
# ---------------------------------------------------------------------------

def compute_sku_metrics(
    order_data: pd.DataFrame,
    sku_mapping: pd.DataFrame,
    store_cohorts: dict,
) -> pd.DataFrame:
    df = order_data.copy()
    df["cohort"] = df["store_id"].map(store_cohorts)
    df["month"] = df["order_date"].dt.to_period("M")

    grouped = df.groupby(["sku", "cohort"]).agg(
        total_units=("total_quantity", "sum"),
        store_penetration=("store_id", "nunique"),
        num_orders=("total_quantity", "count"),
        months_with_orders=("month", "nunique"),
    ).reset_index()

    store_months = df.groupby(["sku", "cohort", "store_id"])["month"].nunique().reset_index(name="store_month_count")
    avg_freq = store_months.groupby(["sku", "cohort"])["store_month_count"].mean().reset_index(name="ordering_frequency")
    avg_freq["ordering_frequency"] = avg_freq["ordering_frequency"] / NUM_MONTHS

    grouped = grouped.merge(avg_freq, on=["sku", "cohort"], how="left")
    grouped["avg_monthly_units"] = grouped["total_units"] / NUM_MONTHS
    grouped["avg_qty_per_order"] = grouped["total_units"] / grouped["num_orders"]
    grouped.drop(columns=["num_orders"], inplace=True)

    grouped = grouped.merge(sku_mapping, on="sku", how="right")
    grouped["no_order_data"] = grouped["cohort"].isna()

    no_data = grouped[grouped["no_order_data"]].drop_duplicates(subset=["sku"])
    has_data = grouped[~grouped["no_order_data"]]

    for col in ["total_units", "store_penetration", "avg_monthly_units", "ordering_frequency", "avg_qty_per_order"]:
        no_data[col] = 0

    return pd.concat([has_data, no_data], ignore_index=True)


# ---------------------------------------------------------------------------
# Computation: store-level summary
# ---------------------------------------------------------------------------

def compute_store_summary(
    order_data: pd.DataFrame,
    rewards_skus: set,
    store_cohorts: dict,
) -> pd.DataFrame:
    df = order_data.copy()
    df["cohort"] = df["store_id"].map(store_cohorts)
    df["is_rewards_sku"] = df["sku"].isin(rewards_skus)

    store_totals = df.groupby("store_id").agg(
        total_coke_units=("total_quantity", "sum"),
        distinct_coke_skus=("sku", "nunique"),
    ).reset_index()

    rewards_df = df[df["is_rewards_sku"]]
    rewards_agg = rewards_df.groupby("store_id").agg(
        total_rewards_units=("total_quantity", "sum"),
        distinct_rewards_skus=("sku", "nunique"),
    ).reset_index()

    non_rewards_agg = df[~df["is_rewards_sku"]].groupby("store_id").agg(
        total_non_rewards_units=("total_quantity", "sum"),
    ).reset_index()

    result = store_totals.merge(rewards_agg, on="store_id", how="left")
    result = result.merge(non_rewards_agg, on="store_id", how="left")
    result["total_rewards_units"] = result["total_rewards_units"].fillna(0).astype(int)
    result["distinct_rewards_skus"] = result["distinct_rewards_skus"].fillna(0).astype(int)
    result["total_non_rewards_units"] = result["total_non_rewards_units"].fillna(0).astype(int)
    result["cohort"] = result["store_id"].map(store_cohorts)

    return result


# ---------------------------------------------------------------------------
# Computation: monthly trend
# ---------------------------------------------------------------------------

def compute_monthly_trend(
    order_data: pd.DataFrame,
    rewards_skus: set,
    store_cohorts: dict,
) -> pd.DataFrame:
    df = order_data.copy()
    df["cohort"] = df["store_id"].map(store_cohorts)
    df["sku_type"] = df["sku"].apply(lambda s: "rewards" if s in rewards_skus else "non-rewards")
    df["month"] = df["order_date"].dt.to_period("M").astype(str)

    result = df.groupby(["month", "cohort", "sku_type"]).agg(
        total_units=("total_quantity", "sum"),
        num_orders=("total_quantity", "count"),
        distinct_stores=("store_id", "nunique"),
        distinct_skus=("sku", "nunique"),
    ).reset_index()

    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("Loading SKU mapping from Excel...")
    sku_mapping = load_sku_mapping()
    print(f"  {len(sku_mapping)} SKUs loaded ({sku_mapping['rewards_match'].notna().sum()} rewards)")

    rewards_skus = set(sku_mapping.loc[sku_mapping["rewards_match"].notna(), "sku"])
    all_skus = list(sku_mapping["sku"])

    order_data = fetch_order_data(all_skus)
    print(f"  {len(order_data):,} order rows, {order_data['store_id'].nunique():,} stores")

    print("Assigning store cohorts...")
    store_cohorts = assign_store_cohorts(order_data, rewards_skus)
    re_count = sum(1 for v in store_cohorts.values() if v == "rewards-eligible")
    co_count = sum(1 for v in store_cohorts.values() if v == "coca-cola-only")
    print(f"  Rewards-eligible: {re_count:,} stores | Coca-Cola only: {co_count:,} stores")

    print("Computing SKU metrics...")
    sku_metrics = compute_sku_metrics(order_data, sku_mapping, store_cohorts)
    sku_metrics.to_csv(OUTPUT_DIR / "sku_monthly_behavior.csv", index=False)
    print(f"  Saved sku_monthly_behavior.csv ({len(sku_metrics)} rows)")

    print("Computing store summaries...")
    store_summary = compute_store_summary(order_data, rewards_skus, store_cohorts)
    store_summary.to_csv(OUTPUT_DIR / "store_cohort_summary.csv", index=False)
    print(f"  Saved store_cohort_summary.csv ({len(store_summary)} rows)")

    print("Computing monthly trends...")
    monthly_trend = compute_monthly_trend(order_data, rewards_skus, store_cohorts)
    monthly_trend.to_csv(OUTPUT_DIR / "monthly_trend.csv", index=False)
    print(f"  Saved monthly_trend.csv ({len(monthly_trend)} rows)")

    print("\nDone. All outputs in M-Rewards/")
