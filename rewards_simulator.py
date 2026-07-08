import altair as alt
import openpyxl
import pandas as pd
import streamlit as st
from pathlib import Path

DATA_DIR = Path(__file__).parent
COCACOLA_EXCEL = DATA_DIR / "M-rewards-cocacola.xlsx"
MONSTER_EXCEL = DATA_DIR / "M-rewards-monster.xlsx"
FERRERA_EXCEL = DATA_DIR / "M-rewards-ferrera.xlsx"
RAW_DATA_PATH = DATA_DIR / "rewards_raw_data.csv"
SKU_BEHAVIOR_PATH = DATA_DIR / "sku_monthly_behavior.csv"

COCACOLA_REWARDS = {
    "8 Dollar Rebate": 5000,
    "Membership 9.99": 10000,
    "Reward 3 (TBD)": 15000,
}

MONSTER_REWARDS = {
    "Reward 1": 6500,
    "Reward 2": 10000,
    "Reward 3": 12000,
    "Reward 4": 15000,
    "Reward 5": 35000,
}

FERRERA_REWARDS = {
    "Reward 1": 5000,
    "Reward 2": 10000,
    "Reward 3": 15000,
}


@st.cache_data
def load_raw_data():
    return pd.read_csv(RAW_DATA_PATH, parse_dates=["order_date"])


@st.cache_data
def load_cocacola_data():
    wb = openpyxl.load_workbook(COCACOLA_EXCEL, read_only=True)
    ims = wb["ims"]
    sku_to_title = {}
    for row in ims.iter_rows(min_row=2, values_only=True):
        if row[3]:
            sku_to_title[row[3]] = row[4]
    wb.close()

    behavior = pd.read_csv(SKU_BEHAVIOR_PATH)
    behavior = behavior[behavior["product_title"] != "Grand Total"]

    sku_records = []
    for _, brow in behavior.iterrows():
        title = brow["product_title"]
        sku = None
        for s, t in sku_to_title.items():
            if t == title:
                sku = s
                break
        sku_records.append({
            "sku": sku,
            "product_title": title,
            "store_penetration": brow["Sum of store_penetration"],
            "current_points": int(brow["Proposed Points"]) if pd.notna(brow["Proposed Points"]) else 0,
        })

    return pd.DataFrame(sku_records)


@st.cache_data
def load_monster_data():
    wb = openpyxl.load_workbook(MONSTER_EXCEL, read_only=True)
    ws = wb["Sheet1"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        proposed = row[7] if row[7] is not None else 0
        records.append({
            "sku": row[0],
            "size": row[1],
            "product_title": row[2],
            "grade": row[3] if row[3] not in (0, None) else "",
            "l30d": row[4] if row[4] not in (0, None) else 0,
            "current_points": int(proposed) if isinstance(proposed, (int, float)) else 0,
        })
    wb.close()
    return pd.DataFrame(records)


@st.cache_data
def load_ferrera_data():
    wb = openpyxl.load_workbook(FERRERA_EXCEL, read_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        points = row[8] if row[8] is not None else 0
        records.append({
            "sku": row[0],
            "brand": row[1],
            "product_title": row[2],
            "category": row[5],
            "current_points": int(points) if isinstance(points, (int, float)) else 0,
        })
    wb.close()
    return pd.DataFrame(records)


@st.cache_data
def compute_store_penetration(raw, skus):
    filtered = raw[raw["sku"].isin(skus)]
    pen = filtered.groupby("sku")["store_id"].nunique().reset_index()
    pen.columns = ["sku", "store_penetration"]
    return pen


def get_month_orders(raw, year, month, skus):
    start = pd.Timestamp(year, month, 1)
    if month == 12:
        end = pd.Timestamp(year + 1, 1, 1)
    else:
        end = pd.Timestamp(year, month + 1, 1)
    filtered = raw[(raw["order_date"] >= start) & (raw["order_date"] < end)]
    filtered = filtered[filtered["sku"].isin(skus)]
    return filtered.groupby(["store_id", "sku"])["total_quantity"].sum().reset_index()


def simulate(month_orders, sku_to_title, points_lookup, reward_thresholds):
    df = month_orders.copy()
    df["product_title"] = df["sku"].map(sku_to_title)
    df["points_per_unit"] = df["product_title"].map(points_lookup).fillna(0)
    df["points_earned"] = df["total_quantity"] * df["points_per_unit"]

    store_points = df.groupby("store_id").agg(
        total_points=("points_earned", "sum"),
        total_units=("total_quantity", "sum"),
        distinct_skus=("sku", "nunique"),
    ).reset_index()

    for reward_name, threshold in reward_thresholds.items():
        store_points[reward_name] = store_points["total_points"] >= threshold

    return store_points


def parse_imported_points(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    elif name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(uploaded_file)
    else:
        return None, "Unsupported file type. Please upload a CSV or Excel file."

    df.columns = df.columns.str.strip().str.lower()
    if "sku" not in df.columns or "points" not in df.columns:
        return None, f"File must contain 'sku' and 'points' columns. Found: {', '.join(df.columns)}"

    df = df[["sku", "points"]].dropna(subset=["sku"])
    df["sku"] = df["sku"].astype(str).str.strip()
    result = {}
    for _, row in df.iterrows():
        try:
            result[str(row["sku"])] = int(float(row["points"]))
        except (ValueError, TypeError):
            result[str(row["sku"])] = 0
    return result, None


def render_import_uploader(prefix, valid_skus, sku_to_title, bulk_state_key):
    uploaded = st.file_uploader(
        "Import proposed points (CSV or Excel with 'sku' and 'points' columns)",
        type=["csv", "xlsx", "xls"],
        key=f"{prefix}_import",
    )
    if uploaded:
        points_map, error = parse_imported_points(uploaded)
        if error:
            st.error(error)
        else:
            matched = {sku: pts for sku, pts in points_map.items() if sku in valid_skus}
            skipped = len(points_map) - len(matched)
            if matched:
                for sku, pts in matched.items():
                    title = sku_to_title.get(sku)
                    if title:
                        st.session_state[bulk_state_key][title] = pts
                st.success(f"Imported points for {len(matched)} SKUs." + (f" {skipped} SKUs skipped (not in this brand)." if skipped else ""))
            else:
                st.warning("No matching SKUs found in the uploaded file.")


def render_reward_thresholds(prefix, defaults):
    st.markdown("---")
    st.subheader("Reward Thresholds")

    state_key = f"{prefix}_rewards_list"
    if state_key not in st.session_state:
        st.session_state[state_key] = list(defaults.items())

    rewards_list = st.session_state[state_key]

    to_remove = None
    if rewards_list:
        cols = st.columns(len(rewards_list))
        thresholds = {}
        for i, (name, default_val) in enumerate(rewards_list):
            with cols[i]:
                thresholds[name] = st.number_input(
                    f"{name} (pts)", min_value=0, value=default_val, step=500, key=f"{prefix}_{name}"
                )
                if st.button("Remove", key=f"{prefix}_remove_{i}", use_container_width=True):
                    to_remove = i
    else:
        thresholds = {}

    if to_remove is not None:
        st.session_state[state_key].pop(to_remove)
        st.rerun()

    add_col1, add_col2, add_col3 = st.columns([2, 2, 1])
    with add_col1:
        new_name = st.text_input("New reward name", value=f"Reward {len(rewards_list) + 1}", key=f"{prefix}_new_name")
    with add_col2:
        new_pts = st.number_input("Points", min_value=0, value=5000, step=500, key=f"{prefix}_new_pts")
    with add_col3:
        st.markdown("<div style='height: 28px'></div>", unsafe_allow_html=True)
        if st.button("Add Reward", key=f"{prefix}_add_reward", use_container_width=True):
            st.session_state[state_key].append((new_name, new_pts))
            st.rerun()

    return thresholds


def render_simulation_results(store_points, reward_thresholds):
    total_stores = len(store_points)

    st.subheader("Simulation Results")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Stores", f"{total_stores:,}")
    m2.metric("Avg Points/Store", f"{store_points['total_points'].mean():,.0f}")
    m3.metric("Median Points/Store", f"{store_points['total_points'].median():,.0f}")
    m4.metric("Max Points/Store", f"{store_points['total_points'].max():,.0f}")

    r_cols = st.columns(len(reward_thresholds))
    for i, (name, threshold) in enumerate(reward_thresholds.items()):
        count = int(store_points[name].sum())
        pct = count / total_stores * 100 if total_stores > 0 else 0
        r_cols[i].metric(name, f"{count:,} stores ({pct:.1f}%)", f"{threshold:,} pts needed")

    st.markdown("---")
    st.subheader("Points Distribution")
    hist_data = store_points[["total_points"]].copy()
    hist_data["total_points"] = hist_data["total_points"].clip(
        upper=hist_data["total_points"].quantile(0.99)
    )

    bars = alt.Chart(hist_data).mark_bar(color="#4A90D9").encode(
        alt.X("total_points:Q", bin=alt.Bin(maxbins=50), title="Points Earned"),
        alt.Y("count()", title="Number of Stores"),
    )
    rules = alt.Chart(pd.DataFrame([
        {"threshold": v, "label": k} for k, v in reward_thresholds.items()
    ])).mark_rule(strokeDash=[5, 5], strokeWidth=2).encode(
        x="threshold:Q",
        color=alt.Color("label:N", title="Reward Threshold"),
    )
    st.altair_chart(bars + rules, use_container_width=True)

    with st.expander("Store-level detail"):
        display_cols = ["store_id", "total_points", "total_units", "distinct_skus"] + list(reward_thresholds.keys())
        st.dataframe(
            store_points[display_cols].sort_values("total_points", ascending=False),
            hide_index=True,
            use_container_width=True,
        )


def render_month_picker(prefix, raw):
    raw["year_month"] = raw["order_date"].dt.to_period("M")
    available = sorted(raw["year_month"].unique())
    labels = [str(p) for p in available]
    default_idx = len(labels) - 1
    selected = st.selectbox("Simulation Month", labels, index=default_idx, key=f"{prefix}_month")
    period = pd.Period(selected)
    return period.year, period.month, selected


def cocacola_page(raw):
    st.title("Coca-Cola — M-Rewards Simulator")

    cocacola_skus_df = load_cocacola_data()
    valid_skus = set(cocacola_skus_df["sku"].dropna())

    year, month, month_label = render_month_picker("coke", raw[raw["sku"].isin(valid_skus)])
    st.caption(f"Simulating with {month_label} ordering data")

    pen = compute_store_penetration(raw, valid_skus)
    cocacola_skus_df = cocacola_skus_df.merge(pen, on="sku", how="left", suffixes=("_orig", ""))
    if "store_penetration_orig" in cocacola_skus_df.columns:
        cocacola_skus_df["store_penetration"] = cocacola_skus_df["store_penetration"].fillna(
            cocacola_skus_df["store_penetration_orig"]
        )
        cocacola_skus_df.drop(columns=["store_penetration_orig"], inplace=True)
    cocacola_skus_df["store_penetration"] = cocacola_skus_df["store_penetration"].fillna(0).astype(int)

    st.subheader("SKU Point Editor")

    if "coke_bulk_points" not in st.session_state:
        st.session_state.coke_bulk_points = {}

    coke_sku_to_title = dict(zip(cocacola_skus_df["sku"], cocacola_skus_df["product_title"]))
    render_import_uploader("coke", valid_skus, coke_sku_to_title, "coke_bulk_points")

    edit_df = cocacola_skus_df[["sku", "product_title", "store_penetration", "current_points"]].copy()
    edit_df.columns = ["SKU", "Product Title", "Store Penetration", "Current Points"]
    edit_df["Proposed Points"] = 0

    for title, pts in st.session_state.coke_bulk_points.items():
        mask = edit_df["Product Title"] == title
        edit_df.loc[mask, "Proposed Points"] = pts

    bulk_col1, bulk_col2, bulk_col3 = st.columns([2, 1, 1])
    with bulk_col1:
        search = st.text_input("Search SKUs", placeholder="Type to filter...", key="coke_search")
    with bulk_col2:
        bulk_value = st.number_input("Bulk point value", min_value=0, max_value=5000, value=100, step=50, key="coke_bulk_val")
    with bulk_col3:
        st.markdown("<div style='height: 28px'></div>", unsafe_allow_html=True)
        apply_bulk = st.button("Apply to selected", use_container_width=True, key="coke_bulk_btn")

    edit_df["Select"] = False
    display_df = edit_df
    if search:
        mask = display_df["Product Title"].str.contains(search, case=False, na=False) | display_df["SKU"].astype(str).str.contains(search, case=False, na=False)
        display_df = display_df[mask]

    col_order = ["Select", "SKU", "Product Title", "Store Penetration", "Current Points", "Proposed Points"]
    edited = st.data_editor(
        display_df[col_order],
        column_config={
            "Select": st.column_config.CheckboxColumn(default=False, width="small"),
            "SKU": st.column_config.TextColumn(disabled=True),
            "Product Title": st.column_config.TextColumn(disabled=True, width="large"),
            "Store Penetration": st.column_config.NumberColumn(disabled=True),
            "Current Points": st.column_config.NumberColumn(disabled=True),
            "Proposed Points": st.column_config.NumberColumn(min_value=0, max_value=5000, step=50),
        },
        hide_index=True,
        use_container_width=True,
        key="coke_sku_editor",
    )

    if apply_bulk:
        selected_titles = edited.loc[edited["Select"] == True, "Product Title"].tolist()
        if selected_titles:
            for title in selected_titles:
                st.session_state.coke_bulk_points[title] = bulk_value
            st.rerun()

    export_df = edited[["SKU", "Product Title", "Store Penetration", "Current Points", "Proposed Points"]].copy()
    st.download_button(
        "Export SKU Data",
        export_df.to_csv(index=False),
        file_name="cocacola_sku_export.csv",
        mime="text/csv",
        key="coke_export",
    )

    points_lookup = dict(zip(edit_df["Product Title"], edit_df["Current Points"]))
    for _, row in edited.iterrows():
        if row["Proposed Points"] > 0:
            points_lookup[row["Product Title"]] = row["Proposed Points"]
    for title, pts in st.session_state.coke_bulk_points.items():
        points_lookup[title] = pts

    reward_thresholds = render_reward_thresholds("coke", COCACOLA_REWARDS)

    st.markdown("---")
    sku_to_title = dict(zip(cocacola_skus_df["sku"], cocacola_skus_df["product_title"]))
    month_orders = get_month_orders(raw, year, month, valid_skus)
    month_orders = month_orders.merge(
        cocacola_skus_df[["sku", "product_title"]], on="sku", how="inner"
    )
    store_points = simulate(month_orders, sku_to_title, points_lookup, reward_thresholds)
    render_simulation_results(store_points, reward_thresholds)


def monster_page(raw):
    st.title("Monster — M-Rewards Simulator")

    monster_skus_df = load_monster_data()
    valid_skus = set(monster_skus_df["sku"].dropna())

    year, month, month_label = render_month_picker("monster", raw[raw["sku"].isin(valid_skus)])
    st.caption(f"Simulating with {month_label} ordering data")

    pen = compute_store_penetration(raw, valid_skus)
    monster_skus_df = monster_skus_df.merge(pen, on="sku", how="left")
    monster_skus_df["store_penetration"] = monster_skus_df["store_penetration"].fillna(0).astype(int)

    st.subheader("SKU Point Editor")

    if "monster_bulk_points" not in st.session_state:
        st.session_state.monster_bulk_points = {}

    monster_sku_to_title = dict(zip(monster_skus_df["sku"], monster_skus_df["product_title"]))
    render_import_uploader("monster", valid_skus, monster_sku_to_title, "monster_bulk_points")

    edit_df = monster_skus_df[["sku", "product_title", "size", "store_penetration", "current_points"]].copy()
    edit_df.columns = ["SKU", "Product Title", "Size", "Store Penetration", "Current Points"]
    edit_df["Proposed Points"] = 0

    for title, pts in st.session_state.monster_bulk_points.items():
        mask = edit_df["Product Title"] == title
        edit_df.loc[mask, "Proposed Points"] = pts

    bulk_col1, bulk_col2, bulk_col3 = st.columns([2, 1, 1])
    with bulk_col1:
        search = st.text_input("Search SKUs", placeholder="Type to filter...", key="monster_search")
    with bulk_col2:
        bulk_value = st.number_input("Bulk point value", min_value=0, max_value=5000, value=100, step=50, key="monster_bulk_val")
    with bulk_col3:
        st.markdown("<div style='height: 28px'></div>", unsafe_allow_html=True)
        apply_bulk = st.button("Apply to selected", use_container_width=True, key="monster_bulk_btn")

    edit_df["Select"] = False
    display_df = edit_df
    if search:
        mask = display_df["Product Title"].str.contains(search, case=False, na=False) | display_df["SKU"].astype(str).str.contains(search, case=False, na=False)
        display_df = display_df[mask]

    col_order = ["Select", "SKU", "Product Title", "Size", "Store Penetration", "Current Points", "Proposed Points"]
    edited = st.data_editor(
        display_df[col_order],
        column_config={
            "Select": st.column_config.CheckboxColumn(default=False, width="small"),
            "SKU": st.column_config.TextColumn(disabled=True),
            "Product Title": st.column_config.TextColumn(disabled=True, width="large"),
            "Size": st.column_config.TextColumn(disabled=True),
            "Store Penetration": st.column_config.NumberColumn(disabled=True),
            "Current Points": st.column_config.NumberColumn(disabled=True),
            "Proposed Points": st.column_config.NumberColumn(min_value=0, max_value=5000, step=50),
        },
        hide_index=True,
        use_container_width=True,
        key="monster_sku_editor",
    )

    if apply_bulk:
        selected_titles = edited.loc[edited["Select"] == True, "Product Title"].tolist()
        if selected_titles:
            for title in selected_titles:
                st.session_state.monster_bulk_points[title] = bulk_value
            st.rerun()

    export_df = edited[["SKU", "Product Title", "Size", "Store Penetration", "Current Points", "Proposed Points"]].copy()
    st.download_button(
        "Export SKU Data",
        export_df.to_csv(index=False),
        file_name="monster_sku_export.csv",
        mime="text/csv",
        key="monster_export",
    )

    points_lookup = dict(zip(edit_df["Product Title"], edit_df["Current Points"]))
    for _, row in edited.iterrows():
        if row["Proposed Points"] > 0:
            points_lookup[row["Product Title"]] = row["Proposed Points"]
    for title, pts in st.session_state.monster_bulk_points.items():
        points_lookup[title] = pts

    reward_thresholds = render_reward_thresholds("monster", MONSTER_REWARDS)

    st.markdown("---")
    sku_to_title = dict(zip(monster_skus_df["sku"], monster_skus_df["product_title"]))
    month_orders = get_month_orders(raw, year, month, valid_skus)
    month_orders = month_orders.merge(
        monster_skus_df[["sku", "product_title"]], on="sku", how="inner"
    )
    store_points = simulate(month_orders, sku_to_title, points_lookup, reward_thresholds)
    render_simulation_results(store_points, reward_thresholds)


def ferrera_page(raw):
    st.title("Ferrera — M-Rewards Simulator")

    ferrera_skus_df = load_ferrera_data()
    valid_skus = set(ferrera_skus_df["sku"].dropna())

    year, month, month_label = render_month_picker("ferrera", raw[raw["sku"].isin(valid_skus)])
    st.caption(f"Simulating with {month_label} ordering data")

    pen = compute_store_penetration(raw, valid_skus)
    ferrera_skus_df = ferrera_skus_df.merge(pen, on="sku", how="left")
    ferrera_skus_df["store_penetration"] = ferrera_skus_df["store_penetration"].fillna(0).astype(int)

    st.subheader("SKU Point Editor")

    if "ferrera_bulk_points" not in st.session_state:
        st.session_state.ferrera_bulk_points = {}

    ferrera_sku_to_title = dict(zip(ferrera_skus_df["sku"], ferrera_skus_df["product_title"]))
    render_import_uploader("ferrera", valid_skus, ferrera_sku_to_title, "ferrera_bulk_points")

    edit_df = ferrera_skus_df[["sku", "product_title", "brand", "category", "store_penetration", "current_points"]].copy()
    edit_df.columns = ["SKU", "Product Title", "Brand", "Category", "Store Penetration", "Current Points"]
    edit_df["Proposed Points"] = 0

    for title, pts in st.session_state.ferrera_bulk_points.items():
        mask = edit_df["Product Title"] == title
        edit_df.loc[mask, "Proposed Points"] = pts

    bulk_col1, bulk_col2, bulk_col3 = st.columns([2, 1, 1])
    with bulk_col1:
        search = st.text_input("Search SKUs", placeholder="Type to filter...", key="ferrera_search")
    with bulk_col2:
        bulk_value = st.number_input("Bulk point value", min_value=0, max_value=5000, value=100, step=50, key="ferrera_bulk_val")
    with bulk_col3:
        st.markdown("<div style='height: 28px'></div>", unsafe_allow_html=True)
        apply_bulk = st.button("Apply to selected", use_container_width=True, key="ferrera_bulk_btn")

    edit_df["Select"] = False
    display_df = edit_df
    if search:
        mask = display_df["Product Title"].str.contains(search, case=False, na=False) | display_df["SKU"].astype(str).str.contains(search, case=False, na=False)
        display_df = display_df[mask]

    col_order = ["Select", "SKU", "Product Title", "Brand", "Category", "Store Penetration", "Current Points", "Proposed Points"]
    edited = st.data_editor(
        display_df[col_order],
        column_config={
            "Select": st.column_config.CheckboxColumn(default=False, width="small"),
            "SKU": st.column_config.TextColumn(disabled=True),
            "Product Title": st.column_config.TextColumn(disabled=True, width="large"),
            "Brand": st.column_config.TextColumn(disabled=True),
            "Category": st.column_config.TextColumn(disabled=True),
            "Store Penetration": st.column_config.NumberColumn(disabled=True),
            "Current Points": st.column_config.NumberColumn(disabled=True),
            "Proposed Points": st.column_config.NumberColumn(min_value=0, max_value=5000, step=50),
        },
        hide_index=True,
        use_container_width=True,
        key="ferrera_sku_editor",
    )

    if apply_bulk:
        selected_titles = edited.loc[edited["Select"] == True, "Product Title"].tolist()
        if selected_titles:
            for title in selected_titles:
                st.session_state.ferrera_bulk_points[title] = bulk_value
            st.rerun()

    export_df = edited[["SKU", "Product Title", "Brand", "Category", "Store Penetration", "Current Points", "Proposed Points"]].copy()
    st.download_button(
        "Export SKU Data",
        export_df.to_csv(index=False),
        file_name="ferrera_sku_export.csv",
        mime="text/csv",
        key="ferrera_export",
    )

    points_lookup = dict(zip(edit_df["Product Title"], edit_df["Current Points"]))
    for _, row in edited.iterrows():
        if row["Proposed Points"] > 0:
            points_lookup[row["Product Title"]] = row["Proposed Points"]
    for title, pts in st.session_state.ferrera_bulk_points.items():
        points_lookup[title] = pts

    reward_thresholds = render_reward_thresholds("ferrera", FERRERA_REWARDS)

    st.markdown("---")
    sku_to_title = dict(zip(ferrera_skus_df["sku"], ferrera_skus_df["product_title"]))
    month_orders = get_month_orders(raw, year, month, valid_skus)
    month_orders = month_orders.merge(
        ferrera_skus_df[["sku", "product_title"]], on="sku", how="inner"
    )
    store_points = simulate(month_orders, sku_to_title, points_lookup, reward_thresholds)
    render_simulation_results(store_points, reward_thresholds)


def main():
    st.set_page_config(page_title="M-Rewards Simulator", layout="wide")

    brand = st.sidebar.radio("Brand", ["Coca-Cola", "Monster", "Ferrera"], index=0)

    raw = load_raw_data()

    if brand == "Coca-Cola":
        cocacola_page(raw)
    elif brand == "Monster":
        monster_page(raw)
    else:
        ferrera_page(raw)


if __name__ == "__main__":
    main()
