"""Shared data access: Excel SKU loaders, Athena pulls, and Railway Postgres."""

from __future__ import annotations

import io
import os
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta

NUM_MONTHS = 6
DATA_DIR = Path(__file__).parent
COCACOLA_EXCEL = DATA_DIR / "M-rewards-cocacola.xlsx"
MONSTER_EXCEL = DATA_DIR / "M-rewards-monster.xlsx"
FERRERA_EXCEL = DATA_DIR / "M-rewards-ferrera.xlsx"

DEFAULT_ATHENA_REGION = "us-west-2"
DEFAULT_ATHENA_S3_STAGING = "s3://mercaso-data-platform-prod/athena/sql/"

ORDERS_COLUMNS = ["store_id", "sku", "order_date", "total_quantity"]

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS orders (
    store_id        text        NOT NULL,
    sku             text        NOT NULL,
    order_date      date        NOT NULL,
    total_quantity  numeric     NOT NULL,
    PRIMARY KEY (store_id, sku, order_date)
);

CREATE TABLE IF NOT EXISTS refresh_state (
    id                    int PRIMARY KEY DEFAULT 1,
    last_refreshed_month  text,
    last_refreshed_at     timestamptz,
    row_count             int
);

CREATE TABLE IF NOT EXISTS brand_proposed_points (
    brand      text        NOT NULL,
    sku        text        NOT NULL,
    points     int         NOT NULL,
    updated_at timestamptz NOT NULL DEFAULT NOW(),
    PRIMARY KEY (brand, sku)
);

CREATE TABLE IF NOT EXISTS brand_rewards (
    brand      text        NOT NULL,
    sort       int         NOT NULL,
    name       text        NOT NULL,
    points     int         NOT NULL,
    PRIMARY KEY (brand, sort)
);

CREATE TABLE IF NOT EXISTS brand_skus (
    brand           text        NOT NULL,
    sku             text        NOT NULL,
    product_title   text        NOT NULL DEFAULT '',
    current_points  int         NOT NULL DEFAULT 0,
    size            text,
    product_brand   text,
    category        text,
    sort            int         NOT NULL DEFAULT 0,
    updated_at      timestamptz NOT NULL DEFAULT NOW(),
    PRIMARY KEY (brand, sku)
);

CREATE TABLE IF NOT EXISTS brands (
    slug       text        PRIMARY KEY,
    label      text        NOT NULL,
    theme      text        NOT NULL DEFAULT 'default',
    sort       int         NOT NULL DEFAULT 0
);
"""

CATALOG_BRANDS = ("coca-cola", "monster", "ferrera")
CATALOG_EXTRA_COLS = ("size", "brand", "category")
CATALOG_CORE_COLS = ("sku", "product_title", "current_points")
PALETTE_THEMES = ("coca-cola", "monster", "ferrera", "default")


class DuplicateBrandError(ValueError):
    """Raised when inserting a brand slug that already exists in the registry."""


# ---------------------------------------------------------------------------
# Date windows
# ---------------------------------------------------------------------------

def previous_month_window(today: date) -> tuple[date, date]:
    """Return [start, end) for the last complete calendar month.

    On September 1, this is August 1 through September 1 (exclusive).
    """
    start = today.replace(day=1) - relativedelta(months=1)
    end = today.replace(day=1)
    return start, end


def backfill_windows(today: date, num_months: int = NUM_MONTHS) -> list[tuple[date, date]]:
    """Return [start, end) windows for the last `num_months` complete months, oldest first."""
    windows = []
    for i in range(num_months, 0, -1):
        start = today.replace(day=1) - relativedelta(months=i)
        end = today.replace(day=1) - relativedelta(months=i - 1)
        windows.append((start, end))
    return windows


def format_month_label(year_month: str) -> str:
    """Turn '2026-08' into 'August 2026'."""
    dt = datetime.strptime(year_month, "%Y-%m")
    return dt.strftime("%B %Y")


# ---------------------------------------------------------------------------
# Excel loaders
# ---------------------------------------------------------------------------

def _safe_int(value, default=0) -> int:
    if value is None or value == "":
        return default
    try:
        if pd.isna(value):
            return default
    except (TypeError, ValueError):
        pass
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _cocacola_ranked_sheet(wb):
    for name in ("sku's-coca-cola", "WOS Ranked"):
        if name in wb.sheetnames:
            return wb[name]
    raise KeyError("No Coca-Cola SKU ranking sheet found (tried sku's-coca-cola, WOS Ranked)")


def load_sku_mapping(excel_path=COCACOLA_EXCEL) -> pd.DataFrame:
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    ims_sheet = wb["ims"]
    sku_to_title = {}
    for row in ims_sheet.iter_rows(min_row=2, values_only=True):
        if row[3]:
            sku_to_title[row[3]] = row[4]

    ws = _cocacola_ranked_sheet(wb)
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = row[2]
        if not sku:
            continue
        records.append({
            "sku": sku,
            "product_title": sku_to_title.get(sku, row[3]),
            "category": row[1],
            "rewards_match": row[11],
            "points": row[10],
        })
    wb.close()
    return pd.DataFrame(records)


def load_cocacola_skus(excel_path=COCACOLA_EXCEL) -> pd.DataFrame:
    mapping = load_sku_mapping(excel_path)
    df = mapping[["sku", "product_title", "points"]].copy()
    df = df.rename(columns={"points": "current_points"})
    df["current_points"] = df["current_points"].map(_safe_int)
    df["sku"] = df["sku"].astype(str)
    return df


def load_monster_skus(excel_path=MONSTER_EXCEL) -> pd.DataFrame:
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb["Sheet1"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        proposed = row[7] if row[7] is not None else 0
        records.append({
            "sku": row[0],
            "size": row[1],
            "product_title": row[2],
            "grade": row[3] if row[3] not in (0, None) else "",
            "l30d": row[4] if row[4] not in (0, None) else 0,
            "current_points": _safe_int(proposed),
        })
    wb.close()
    return pd.DataFrame(records)


def load_ferrera_skus(excel_path=FERRERA_EXCEL) -> pd.DataFrame:
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        points = row[8] if row[8] is not None else 0
        records.append({
            "sku": row[0],
            "brand": row[1],
            "product_title": row[2],
            "category": row[5],
            "current_points": _safe_int(points),
        })
    wb.close()
    return pd.DataFrame(records)


def excel_skus_for_brand(brand: str) -> pd.DataFrame:
    if brand == "coca-cola":
        return load_cocacola_skus()
    if brand == "monster":
        return load_monster_skus()
    if brand == "ferrera":
        return load_ferrera_skus()
    raise KeyError(brand)


# ---------------------------------------------------------------------------
# Catalog records (canonical SKU list)
# ---------------------------------------------------------------------------

def normalize_catalog_row(row: dict) -> dict:
    sku = str(row.get("sku") or "").strip()
    title_raw = row.get("product_title")
    try:
        title_missing = title_raw is None or pd.isna(title_raw)
    except (TypeError, ValueError):
        title_missing = title_raw is None
    title = "" if title_missing else str(title_raw).strip()
    if title.lower() == "nan":
        title = ""
    rec = {
        "sku": sku,
        "product_title": title,
        "current_points": _safe_int(row.get("current_points", 0)),
    }
    for col in CATALOG_EXTRA_COLS:
        val = row.get(col)
        if val is None:
            continue
        try:
            if pd.isna(val):
                continue
        except (TypeError, ValueError):
            pass
        text = str(val).strip()
        if text and text.lower() != "nan":
            rec[col] = text
    return rec


def records_from_sku_frame(df: pd.DataFrame) -> list[dict]:
    if df is None or df.empty:
        return []
    records = []
    seen = {}
    for _, row in df.iterrows():
        rec = normalize_catalog_row(row.to_dict())
        if not rec["sku"] or rec["sku"].lower() == "nan":
            continue
        seen[rec["sku"]] = rec
    for rec in seen.values():
        records.append(rec)
    return records


def catalog_frame(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame(columns=list(CATALOG_CORE_COLS))
    normalized = [normalize_catalog_row(r) for r in records if str(r.get("sku") or "").strip()]
    df = pd.DataFrame.from_records(normalized)
    df["sku"] = df["sku"].astype(str)
    df["product_title"] = df["product_title"].fillna("").astype(str)
    df["current_points"] = df["current_points"].map(_safe_int)
    return df


def overlay_catalog(existing: list[dict], incoming: list[dict]) -> list[dict]:
    """Merge incoming SKUs onto existing. Omitted SKUs are kept; incoming order for new SKUs."""
    existing_n = [normalize_catalog_row(r) for r in existing if str(r.get("sku") or "").strip()]
    incoming_n = [normalize_catalog_row(r) for r in incoming if str(r.get("sku") or "").strip()]
    incoming_by = {r["sku"]: r for r in incoming_n}
    out = []
    seen = set()
    for row in existing_n:
        sku = row["sku"]
        out.append(incoming_by.get(sku, row))
        seen.add(sku)
    for row in incoming_n:
        if row["sku"] not in seen:
            out.append(row)
            seen.add(row["sku"])
    return out


def diff_catalog(existing: list[dict], incoming: list[dict]) -> dict:
    existing_n = [normalize_catalog_row(r) for r in existing if str(r.get("sku") or "").strip()]
    incoming_n = [normalize_catalog_row(r) for r in incoming if str(r.get("sku") or "").strip()]
    existing_by = {r["sku"]: r for r in existing_n}
    incoming_by = {r["sku"]: r for r in incoming_n}
    added = [incoming_by[s] for s in incoming_by if s not in existing_by]
    removed = [existing_by[s] for s in existing_by if s not in incoming_by]
    updated = []
    for sku, new in incoming_by.items():
        old = existing_by.get(sku)
        if old is None:
            continue
        if old != new:
            updated.append({"sku": sku, "before": old, "after": new})
    return {
        "added": added,
        "removed": removed,
        "updated": updated,
        "unchanged": len(incoming_by) - len(added) - len(updated),
        "incoming_count": len(incoming_n),
        "existing_count": len(existing_n),
    }


def parse_catalog_file(file_bytes: bytes, filename: str) -> tuple[list[dict] | None, str | None]:
    """Parse a canonical catalog CSV/Excel (sku, product_title, current_points; extras optional)."""
    name = (filename or "").lower()
    buffer = io.BytesIO(file_bytes)
    try:
        if name.endswith(".csv"):
            df = pd.read_csv(buffer)
        elif name.endswith((".xlsx", ".xls")):
            df = pd.read_excel(buffer)
        else:
            return None, "Unsupported file type. Please upload a CSV or Excel file."
    except Exception as exc:
        return None, f"Could not read file: {exc}"

    if df is None or df.empty:
        return None, "No SKU rows found in the uploaded file."

    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    aliases = {
        "title": "product_title",
        "product": "product_title",
        "points": "current_points",
        "current": "current_points",
    }
    df = df.rename(columns={k: v for k, v in aliases.items() if k in df.columns})
    if "sku" not in df.columns or "product_title" not in df.columns:
        return None, (
            "File must contain 'sku' and 'product_title' columns "
            "(optional: current_points, size, brand, category). "
            f"Found: {', '.join(df.columns)}"
        )
    if "current_points" not in df.columns:
        df["current_points"] = 0
    records = records_from_sku_frame(df)
    if not records:
        return None, "No SKU rows found in the uploaded file."
    return records, None


def catalog_download_frame(records: list[dict]) -> pd.DataFrame:
    df = catalog_frame(records)
    cols = list(CATALOG_CORE_COLS)
    for col in CATALOG_EXTRA_COLS:
        if col in df.columns and df[col].notna().any():
            nonempty = df[col].astype(str).str.strip()
            if (nonempty != "").any() and (nonempty.str.lower() != "nan").any():
                cols.append(col)
    cols = [c for c in cols if c in df.columns]
    return df[cols]


def load_all_skus(conn=None) -> list[str]:
    """SKU union for Athena refresh: Postgres catalogs, seeded from Excel if empty."""
    conn, close = _borrow_connection(conn)
    try:
        seed_brand_catalogs(conn)
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT sku FROM brand_skus ORDER BY sku")
            return [str(row[0]) for row in cur.fetchall()]
    finally:
        if close:
            conn.close()


# ---------------------------------------------------------------------------
# Athena
# ---------------------------------------------------------------------------

def _sql_quote(value) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def athena_config() -> tuple[str, str]:
    region = os.environ.get("AWS_DEFAULT_REGION") or os.environ.get("ATHENA_REGION") or DEFAULT_ATHENA_REGION
    staging = os.environ.get("ATHENA_S3_STAGING") or DEFAULT_ATHENA_S3_STAGING
    return region, staging


def _athena_connect():
    from pyathena import connect

    region, staging = athena_config()
    return connect(s3_staging_dir=staging, region_name=region)


def empty_orders_frame() -> pd.DataFrame:
    return pd.DataFrame(columns=ORDERS_COLUMNS)


def fetch_order_data(sku_list, start: date, end: date, connect_fn=None) -> pd.DataFrame:
    """Query Athena for store/SKU quantities in [start, end).

    Inclusion: any store that ordered a catalog SKU in the window, with no
    min-quantity HAVING filter, cancelled-order predicate, or store-status
    filter. Quantity is SUM(li.initial_quantity), not net of returns.
    """
    if not sku_list:
        return empty_orders_frame()

    sku_values = ", ".join(_sql_quote(s) for s in sku_list)
    start_ts = f"{start.isoformat()} 00:00:00"
    end_ts = f"{end.isoformat()} 00:00:00"
    query = f"""
    SELECT
        o.store_id,
        li.sku,
        DATE(li.order_item_created_at) AS order_date,
        SUM(li.initial_quantity)       AS total_quantity
    FROM dwm.dwm_trade_line_item_detail_full li
    JOIN dwm.dwm_trade_order_detail_full o
        ON li.order_id = o.order_id
       AND o.dt = (SELECT MAX(dt) FROM dwm.dwm_trade_order_detail_full)
    WHERE li.dt = (SELECT MAX(dt) FROM dwm.dwm_trade_line_item_detail_full)
      AND li.order_item_created_at >= TIMESTAMP '{start_ts}'
      AND li.order_item_created_at <  TIMESTAMP '{end_ts}'
      AND li.sku IN ({sku_values})
    GROUP BY o.store_id, li.sku, DATE(li.order_item_created_at)
    ORDER BY o.store_id, li.sku, DATE(li.order_item_created_at)
    """

    connect_fn = connect_fn or _athena_connect
    conn = connect_fn()
    df = pd.read_sql(query, conn)
    if df is None or df.empty:
        return empty_orders_frame()
    df["order_date"] = pd.to_datetime(df["order_date"])
    df["store_id"] = df["store_id"].astype(str)
    df["sku"] = df["sku"].astype(str)
    df["total_quantity"] = pd.to_numeric(df["total_quantity"])
    return df[ORDERS_COLUMNS]


# ---------------------------------------------------------------------------
# Postgres
# ---------------------------------------------------------------------------

def get_database_url() -> str:
    """Resolve Postgres URL from Railway-style env vars."""
    for key in ("DATABASE_URL", "DATABASE_PRIVATE_URL"):
        url = os.environ.get(key, "").strip()
        if url and not url.startswith("${{"):
            return url

    host = os.environ.get("PGHOST", "").strip()
    port = os.environ.get("PGPORT", "5432").strip()
    user = os.environ.get("PGUSER", "").strip()
    password = os.environ.get("PGPASSWORD", "")
    database = os.environ.get("PGDATABASE", "").strip()

    if host and user and password and database:
        from urllib.parse import quote_plus

        return (
            f"postgresql://{quote_plus(user)}:{quote_plus(password)}"
            f"@{host}:{port}/{quote_plus(database)}"
        )

    present = [k for k in (
        "DATABASE_URL", "DATABASE_PRIVATE_URL",
        "PGHOST", "PGPORT", "PGUSER", "PGPASSWORD", "PGDATABASE",
    ) if os.environ.get(k)]

    hint = (
        "On the **web** service Variables tab, add a reference: "
        "`DATABASE_URL` → `${{YourPostgresServiceName.DATABASE_URL}}` "
        "(use your Postgres service's exact name). Then redeploy web."
    )
    if present:
        raise RuntimeError(
            "Could not build a Postgres connection string. "
            f"Found {', '.join(present)} but not a usable DATABASE_URL. {hint}"
        )
    raise RuntimeError(
        "DATABASE_URL is not set. Point the **web** service at Railway Postgres, "
        f"then redeploy. {hint}"
    )

def get_connection(database_url: str | None = None):
    import psycopg2

    return psycopg2.connect(database_url or get_database_url())


def init_schema(conn) -> None:
    statements = [s.strip() for s in SCHEMA_SQL.split(";") if s.strip()]
    with conn.cursor() as cur:
        for statement in statements:
            cur.execute(statement)
    conn.commit()


def ensure_schema() -> bool:
    """Create tables when DATABASE_URL is configured. Returns True if init ran.

    Used by the web app on startup so new tables exist without waiting for cron.
    Unit tests (no database URL) skip this and get False.
    """
    try:
        get_database_url()
    except RuntimeError:
        return False
    conn = get_connection()
    try:
        init_schema(conn)
        seed_brand_registry(conn)
        seed_brand_catalogs(conn)
        return True
    finally:
        conn.close()


def _borrow_connection(conn):
    """Return (conn, should_close). Opens a new connection when conn is None."""
    if conn is not None:
        return conn, False
    return get_connection(), True


def overlay_proposed_points(existing: dict[str, int], patch: dict[str, int]) -> dict[str, int]:
    """Merge a proposed-points patch onto an existing map.

    Keys omitted from `patch` are kept. Values ≤ 0 remove that SKU's override
    so catalog current_points apply again.
    """
    out = {str(k): int(v) for k, v in existing.items() if int(v) > 0}
    for sku, pts in patch.items():
        sku = str(sku)
        try:
            value = int(pts)
        except (TypeError, ValueError):
            value = 0
        if value > 0:
            out[sku] = value
        else:
            out.pop(sku, None)
    return out


def load_proposed_points(brand: str, conn=None) -> dict[str, int]:
    conn, close = _borrow_connection(conn)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT sku, points FROM brand_proposed_points WHERE brand = %s",
                (brand,),
            )
            rows = cur.fetchall()
        return {str(sku): int(pts) for sku, pts in rows if int(pts) > 0}
    finally:
        if close:
            conn.close()


def replace_proposed_points(brand: str, mapping: dict[str, int], conn=None) -> None:
    from psycopg2.extras import execute_values

    conn, close = _borrow_connection(conn)
    try:
        rows = [
            (brand, str(sku), int(pts))
            for sku, pts in mapping.items()
            if int(pts) > 0
        ]
        with conn.cursor() as cur:
            cur.execute("DELETE FROM brand_proposed_points WHERE brand = %s", (brand,))
            if rows:
                execute_values(
                    cur,
                    "INSERT INTO brand_proposed_points (brand, sku, points) VALUES %s",
                    rows,
                    page_size=1000,
                )
        conn.commit()
    finally:
        if close:
            conn.close()


def merge_proposed_points(brand: str, patch: dict[str, int], conn=None) -> dict[str, int]:
    """Load existing overrides, overlay `patch`, write back. Returns the merged map."""
    conn, close = _borrow_connection(conn)
    try:
        existing = load_proposed_points(brand, conn=conn)
        merged = overlay_proposed_points(existing, patch)
        replace_proposed_points(brand, merged, conn=conn)
        return merged
    finally:
        if close:
            conn.close()


def load_brand_rewards(brand: str, conn=None) -> list[tuple[str, int]]:
    conn, close = _borrow_connection(conn)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT name, points FROM brand_rewards WHERE brand = %s ORDER BY sort",
                (brand,),
            )
            rows = cur.fetchall()
        return [(str(name), int(pts)) for name, pts in rows]
    finally:
        if close:
            conn.close()


def save_brand_rewards(brand: str, rewards: list[tuple[str, int]], conn=None, commit: bool = True) -> None:
    from psycopg2.extras import execute_values

    conn, close = _borrow_connection(conn)
    try:
        rows = [
            (brand, idx, str(name), int(pts))
            for idx, (name, pts) in enumerate(rewards)
        ]
        with conn.cursor() as cur:
            cur.execute("DELETE FROM brand_rewards WHERE brand = %s", (brand,))
            if rows:
                execute_values(
                    cur,
                    "INSERT INTO brand_rewards (brand, sort, name, points) VALUES %s",
                    rows,
                    page_size=100,
                )
        if commit:
            conn.commit()
    finally:
        if close:
            conn.close()


def count_catalog_skus(brand: str, conn=None) -> int:
    conn, close = _borrow_connection(conn)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM brand_skus WHERE brand = %s", (brand,))
            return int(cur.fetchone()[0])
    finally:
        if close:
            conn.close()


def load_catalog_skus(brand: str, conn=None) -> list[dict]:
    conn, close = _borrow_connection(conn)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT sku, product_title, current_points, size, product_brand, category "
                "FROM brand_skus WHERE brand = %s ORDER BY sort, sku",
                (brand,),
            )
            rows = cur.fetchall()
        records = []
        for sku, title, points, size, product_brand, category in rows:
            rec = {
                "sku": str(sku),
                "product_title": "" if title is None else str(title),
                "current_points": _safe_int(points),
            }
            if size:
                rec["size"] = str(size)
            if product_brand:
                rec["brand"] = str(product_brand)
            if category:
                rec["category"] = str(category)
            records.append(rec)
        return records
    finally:
        if close:
            conn.close()


def replace_catalog_skus(brand: str, records: list[dict], conn=None, commit: bool = True) -> int:
    from psycopg2.extras import execute_values

    conn, close = _borrow_connection(conn)
    try:
        normalized = [
            normalize_catalog_row(r)
            for r in records
            if str(r.get("sku") or "").strip()
        ]
        keep = [r["sku"] for r in normalized]
        rows = []
        for idx, rec in enumerate(normalized):
            rows.append((
                brand,
                rec["sku"],
                rec.get("product_title") or "",
                _safe_int(rec.get("current_points")),
                rec.get("size"),
                rec.get("brand"),
                rec.get("category"),
                idx,
            ))
        with conn.cursor() as cur:
            cur.execute("DELETE FROM brand_skus WHERE brand = %s", (brand,))
            if rows:
                execute_values(
                    cur,
                    "INSERT INTO brand_skus "
                    "(brand, sku, product_title, current_points, size, product_brand, category, sort) "
                    "VALUES %s",
                    rows,
                    page_size=500,
                )
            if keep:
                cur.execute(
                    "DELETE FROM brand_proposed_points "
                    "WHERE brand = %s AND NOT (sku = ANY(%s))",
                    (brand, keep),
                )
            else:
                cur.execute(
                    "DELETE FROM brand_proposed_points WHERE brand = %s",
                    (brand,),
                )
        if commit:
            conn.commit()
        return len(normalized)
    finally:
        if close:
            conn.close()


def merge_catalog_skus(brand: str, records: list[dict], conn=None) -> list[dict]:
    conn, close = _borrow_connection(conn)
    try:
        existing = load_catalog_skus(brand, conn=conn)
        merged = overlay_catalog(existing, records)
        replace_catalog_skus(brand, merged, conn=conn)
        return merged
    finally:
        if close:
            conn.close()


def load_brands(conn=None) -> list[dict]:
    conn, close = _borrow_connection(conn)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT slug, label, theme, sort FROM brands ORDER BY sort, slug")
            rows = cur.fetchall()
        return [
            {
                "slug": str(slug),
                "label": str(label),
                "theme": str(theme or "default"),
                "sort": int(sort),
            }
            for slug, label, theme, sort in rows
        ]
    finally:
        if close:
            conn.close()


def get_brand(slug: str, conn=None) -> dict | None:
    conn, close = _borrow_connection(conn)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT slug, label, theme, sort FROM brands WHERE slug = %s",
                (slug,),
            )
            row = cur.fetchone()
        if not row:
            return None
        slug, label, theme, sort = row
        return {
            "slug": str(slug),
            "label": str(label),
            "theme": str(theme or "default"),
            "sort": int(sort),
        }
    finally:
        if close:
            conn.close()


def seed_brand_registry(conn=None) -> int:
    """Insert Coca-Cola / Monster / Ferrera if the brands table is empty."""
    from psycopg2.extras import execute_values

    from simulator import BRAND_DEFAULTS

    conn, close = _borrow_connection(conn)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM brands")
            if int(cur.fetchone()[0]) > 0:
                return 0
            rows = [
                (slug, meta["label"], slug, idx)
                for idx, (slug, meta) in enumerate(BRAND_DEFAULTS.items())
            ]
            execute_values(
                cur,
                "INSERT INTO brands (slug, label, theme, sort) VALUES %s",
                rows,
            )
        conn.commit()
        return len(rows)
    finally:
        if close:
            conn.close()


def create_brand(
    slug: str,
    label: str,
    theme: str,
    records: list[dict],
    rewards: list[tuple[str, int]],
    conn=None,
) -> dict:
    """Insert a registry row plus catalog and rewards in one transaction."""
    conn, close = _borrow_connection(conn)
    try:
        existing = get_brand(slug, conn=conn)
        if existing:
            raise DuplicateBrandError(slug)
        with conn.cursor() as cur:
            cur.execute("SELECT COALESCE(MAX(sort), -1) + 1 FROM brands")
            sort = int(cur.fetchone()[0])
            cur.execute(
                "INSERT INTO brands (slug, label, theme, sort) VALUES (%s, %s, %s, %s)",
                (slug, label, theme, sort),
            )
        replace_catalog_skus(slug, records, conn=conn, commit=False)
        save_brand_rewards(slug, rewards, conn=conn, commit=False)
        conn.commit()
        return {"slug": slug, "label": label, "theme": theme, "sort": sort}
    except Exception:
        conn.rollback()
        raise
    finally:
        if close:
            conn.close()


def seed_brand_catalogs(conn=None) -> dict[str, int]:
    """Copy git Excel workbooks into Postgres for brands with an empty catalog."""
    conn, close = _borrow_connection(conn)
    try:
        seeded = {}
        for brand in CATALOG_BRANDS:
            if count_catalog_skus(brand, conn=conn) > 0:
                continue
            records = records_from_sku_frame(excel_skus_for_brand(brand))
            if not records:
                continue
            replace_catalog_skus(brand, records, conn=conn)
            seeded[brand] = len(records)
        return seeded
    finally:
        if close:
            conn.close()


def get_refresh_state(conn) -> dict | None:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT last_refreshed_month, last_refreshed_at, row_count "
            "FROM refresh_state WHERE id = 1"
        )
        row = cur.fetchone()
    if not row:
        return None
    return {
        "last_refreshed_month": row[0],
        "last_refreshed_at": row[1],
        "row_count": row[2],
    }


def set_refresh_state(conn, month: str, row_count: int) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO refresh_state (id, last_refreshed_month, last_refreshed_at, row_count)
            VALUES (1, %s, NOW(), %s)
            ON CONFLICT (id) DO UPDATE SET
                last_refreshed_month = EXCLUDED.last_refreshed_month,
                last_refreshed_at = EXCLUDED.last_refreshed_at,
                row_count = EXCLUDED.row_count
            """,
            (month, row_count),
        )
    conn.commit()


def replace_month_orders_frame(
    existing: pd.DataFrame,
    incoming: pd.DataFrame,
    start: date,
    end: date,
) -> pd.DataFrame:
    """Replace rows in [start, end) with incoming. Pure DataFrame helper for tests."""
    if existing is None or existing.empty:
        kept = empty_orders_frame()
    else:
        kept = existing.copy()
        kept["order_date"] = pd.to_datetime(kept["order_date"])
        start_ts = pd.Timestamp(start)
        end_ts = pd.Timestamp(end)
        kept = kept[(kept["order_date"] < start_ts) | (kept["order_date"] >= end_ts)]
    if incoming is None or incoming.empty:
        new_rows = empty_orders_frame()
    else:
        new_rows = incoming.copy()
        new_rows["order_date"] = pd.to_datetime(new_rows["order_date"])
    return pd.concat([kept, new_rows], ignore_index=True)


def _normalize_order_date(value) -> date:
    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return pd.to_datetime(value).date()


def _order_rows(orders: pd.DataFrame) -> list[tuple]:
    rows = []
    for rec in orders.itertuples(index=False):
        rows.append((
            str(rec.store_id),
            str(rec.sku),
            _normalize_order_date(rec.order_date),
            float(rec.total_quantity),
        ))
    return rows


def replace_month_orders(conn, orders: pd.DataFrame, start: date, end: date) -> int:
    """Delete existing rows in [start, end) and insert `orders`. Returns inserted count."""
    from psycopg2.extras import execute_values

    with conn.cursor() as cur:
        cur.execute(
            "DELETE FROM orders WHERE order_date >= %s AND order_date < %s",
            (start, end),
        )
        if orders is None or orders.empty:
            conn.commit()
            return 0
        rows = _order_rows(orders)
        execute_values(
            cur,
            "INSERT INTO orders (store_id, sku, order_date, total_quantity) VALUES %s",
            rows,
            page_size=1000,
        )
    conn.commit()
    return len(rows)


def count_orders(conn) -> int:
    with conn.cursor() as cur:
        cur.execute("SELECT COUNT(*) FROM orders")
        return int(cur.fetchone()[0])


def read_orders(conn=None) -> pd.DataFrame:
    close = False
    if conn is None:
        conn = get_connection()
        close = True
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT store_id, sku, order_date, total_quantity FROM orders"
            )
            rows = cur.fetchall()
        df = pd.DataFrame(rows, columns=ORDERS_COLUMNS)
        if df.empty:
            return empty_orders_frame()
        df["order_date"] = pd.to_datetime(df["order_date"])
        df["store_id"] = df["store_id"].astype(str)
        df["sku"] = df["sku"].astype(str)
        df["total_quantity"] = pd.to_numeric(df["total_quantity"])
        return df
    finally:
        if close:
            conn.close()
